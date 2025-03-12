from kor.extraction import create_extraction_chain
from kor.nodes import Object, Text, Number

# LangChain Models
from langchain_core.prompts import PromptTemplate
from langchain_openai import OpenAI
from langchain_openai import ChatOpenAI

# Standard Helpers
import time
import json
from datetime import datetime
import os
import openpyxl

# For token counting
from langchain.callbacks import get_openai_callback

gpt = "gpt-4o-mini"

# Fetch API key from environment
api_key = os.getenv("api_key")
if not api_key:
    raise ValueError("API key is required but not set. Use os.environ['api_key'] = 'your-key' before running.")

llm = ChatOpenAI(
    model_name=gpt,
    temperature=0,
    max_tokens=2000,
    openai_api_key=api_key
)

# Allow instruction to be overridden dynamically
def get_chain(instruction=None):
    """Create an extraction chain, allowing instruction override."""
    if instruction is None:
        # Default to an empty Object node if no instruction is provided
        instruction = Object(
            id="default",
            description="Default extraction object",
            attributes=[]  # No attributes by default
        )
    return create_extraction_chain(llm, instruction, encoder_or_encoder_class="csv")

# Initialize with a default instruction
chain = get_chain()

def load_processed_files(log_file):
    """Load the list of processed files from the log file."""
    if os.path.exists(log_file):
        with open(log_file, 'r') as f:
            return set(line.strip() for line in f)
    return set()

def update_processed_files(log_file, processed_files):
    """Update the log file with the list of processed files."""
    with open(log_file, 'a') as f:
        for filename in processed_files:
            f.write(f"{filename}\n")

def handle_value(value):
    """Convert value to a string and handle lists or empty values."""
    if isinstance(value, list):
        return ', '.join(map(str, value))
    return value if value else "NA"

def pasqui_structuring(summaries_out, results_file, errors_file, log_file, headers_vars, instruction=None):
    """Process text files and structure results into an Excel file."""

    # Load already processed files
    processed_files = load_processed_files(log_file)

    # Check if the workbook already exists
    if os.path.exists(results_file):
        wb = openpyxl.load_workbook(results_file)
        results_sheet = wb["Results"]
        errors_sheet = wb["Errors"]
    else:
        wb = openpyxl.Workbook()
        results_sheet = wb.active
        results_sheet.title = "Results"
        errors_sheet = wb.create_sheet(title="Errors")
        results_sheet.append(headers_vars)
        errors_sheet.append(["Error Files"])

    error_files = []
    new_processed_files = set()

    # Get and sort text files
    files = sorted([f for f in os.listdir(summaries_out) if f.endswith(".txt")])

    # Create a new chain if an instruction override is provided
    local_chain = get_chain(instruction) if instruction else chain

    for filename in files:
        if filename not in processed_files:
            file_path = os.path.join(summaries_out, filename)
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
                input_data = {"text": text}

                try:
                    # Process with LangChain extraction
                    output = local_chain.invoke(input_data)
                    print(f"Output for {filename}: {output}")

                    data = output.get('data', {})
                    instruction_list = data.get("instruction", [])

                    if instruction_list:
                        for instruction in instruction_list:
                            row = [filename]
                            for header in headers_vars[1:]:
                                value = instruction.get(header)
                                row.append(handle_value(value))
                            results_sheet.append(row)

                        # Debugging: Print rows after appending to the sheet
                        print("Data being written to sheet:")
                        for row in results_sheet.iter_rows(values_only=True):
                            print(row)  # Debugging to see what has been appended

                        wb.save(results_file)  # Save after every iteration

                    else:
                        results_sheet.append([filename] + ["NA"] * (len(headers_vars) - 1))

                    new_processed_files.add(filename)
                    wb.save(results_file)
                    update_processed_files(log_file, {filename})

                except Exception as e:
                    print(f"Error processing file {filename}: {e}")
                    error_files.append([filename])

    for filename in error_files:
        errors_sheet.append(filename)

    wb.save(results_file)
    update_processed_files(log_file, new_processed_files)
    
    print(f"Results saved to {results_file}")
    print(f"Errors saved to {errors_file}")

